# pdf_scraper.py

import sys
import PyPDF2
import nltk.data
from openpyxl import load_workbook

keywords = ['shall', 'must', 'may', 'could', 'can', 'should', 'option', 'will', 'discretion', 'sole', 'required', 'remedy', 'minimum', 'maximum', 'pay', 'remit']

# workbook
file = "C:/Users/tyler/OneDrive/Documents/Personal/Coding/pdf-scraper/Sentences.xlsx"
wb = load_workbook(filename=file)
oRow = 2

# creating an object 
pdfFile = open("C:/Users/tyler/OneDrive/Documents/Personal/Coding/pdf-scraper/Lockheed_Martin.pdf", 'rb')
text = PyPDF2.PdfFileReader(pdfFile)
numPages = text.getNumPages()
page_text = ""
for i in range(numPages):
    page_text += text.getPage(i).extractText()

# split sentences
tokenizer = nltk.data.load('tokenizers/punkt/english.pickle')
sentences = tokenizer.tokenize(page_text)
# nltk folder: C:\Users\tyler\AppData\Roaming\nltk_data

print("Analyzing sentences")
for sentence in sentences:
    sentence = sentence.replace("\n", "")
    sentencePasted = False
    for keyword in keywords:
        if (keyword in sentence) and (sentencePasted == False):
            # output to excel sheet
            wb["Sheet1"].cell(row = oRow, column = 1).value = sentence
            oRow += 1
            sentencePasted = True

# save sheet
try:
    wb.save(filename=file)
except:
    print("Excel workbook is open. Please close the workbook and try again.")
    sys.exit()

print("Program complete")